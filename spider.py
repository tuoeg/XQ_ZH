import requests
import json
import time
from openpyxl import workbook,Workbook
a=[["组合名","组合代码","创建时间","更新时间"]]#全局
histroy_data_set=[]#暂存提取的历史数据
data={'1':9}
class spider(object):
    # def __init__(self) -> None:
    #     super().__init__()
    def request_data(self,url,header):
        req=requests.request('GET',url=url,headers=header)#返回json数据
        print("编码",req.encoding)
        #对json数据进行存储
        print("请求状态码:{0}".format(req.status_code))
        # with open("orignal_json.txt","w",encoding='utf-8') as f:
        #     f.write(req.text)
        return req.text
    def get_effective_data(symbol):
        #histroy_data_set=[]#暂存提取的历史数据
        url='https://xueqiu.com/cubes/rebalancing/history.json?cube_symbol='+str(symbol)+'&count=5&page=1'
        history_json_data=requests.request('GET',url=url,headers=header)
        #ws2=wb.create_sheet('history_data_page',0)
        # with open(str(symbol)+"_data.txt","w",encoding='utf-8') as f:
        #     f.write(history_json_data.text) 
        history_data=json.loads(history_json_data.text)
        for i in range(len(history_data["list"])):
            if history_data["list"][i]["status"]=='failed':
                continue
            for j in range(len(history_data["list"][i]["rebalancing_histories"])):
                stock_symbol=history_data["list"][i]["rebalancing_histories"][j]["stock_symbol"]
                stock_name=history_data["list"][i]["rebalancing_histories"][j]["stock_name"]
                prev_weight=history_data["list"][i]["rebalancing_histories"][j]["prev_weight_adjusted"]
                target_weight=history_data["list"][i]["rebalancing_histories"][j]["target_weight"]
                upate_time=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(history_data["list"][i]["rebalancing_histories"][j]["updated_at"]/1000))
                histroy_data_set.append([symbol,stock_symbol,stock_name,upate_time,prev_weight,target_weight])
    def json_parse(self,json_data):
        #对json数据进行解析，并存取组合代码以及
        # with open('./orignal_json.txt','r',encoding='utf-8') as f:
        #     json_data=f.read()
        a=[["组合名","组合代码","创建时间","更新时间"]]#全局
        json_data=json.loads(json_data)
        #print(json_data["list"][0]['name'])
        #print(len(json_data["list"]))
        for i in range(len(json_data["list"])):
            #时间戳转换
            begin_date=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(json_data["list"][i]['created_at']/1000))
            update_date=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(json_data["list"][i]['updated_at']/1000))
            a.append([str(json_data["list"][i]['name']),str(json_data["list"][i]['symbol']),str(begin_date),str(update_date)])
            print(a)
            spider.get_effective_data(json_data["list"][i]['symbol'])
            time.sleep(1.5)#加上延时，请求过快会被拒绝
            #break
        # for j in a:
        #     ws1.append(j)
        #get_effective_data()
        #ws1.append(a)
        #print(a)
        #print(json_data)
        #print()
    #对1000组合进行爬取，并进行分析
    def data_analysis(self,histroy_data):
        for i in histroy_data:
            #print(i[2])
            if i[2] not in data:
                data[i[2]]=1
            if i[2] in data:
                data[i[2]]=data[i[2]]+1
header={
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36',
            'cookie':'remember=1; xq_is_login=1; u=1080515058; s=br17esaf77; bid=9b5a9df3ebe0a8c95ba95ff4bfd0fa58_kqxiugtb; __utmz=1.1625906667.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); Hm_lvt_fe218c11eab60b6ab1b6f84fb38bcc4a=1625970637; xq_a_token=dba3dd3cb6b6b007fc3c7bcdce62761fe2095e82; xqat=dba3dd3cb6b6b007fc3c7bcdce62761fe2095e82; xq_id_token=eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiJ9.eyJ1aWQiOjEwODA1MTUwNTgsImlzcyI6InVjIiwiZXhwIjoxNjMwOTgwMzA2LCJjdG0iOjE2MjgzODgzMDY0NjgsImNpZCI6ImQ5ZDBuNEFadXAifQ.BYTBUDtIuuPrCFF0P4v0YlEF91qh-3GE_ddO8rFAcOF_DN_aZEEx0MGoeI_mLRFWM08kTLWrVInCdNWR3ImerPr5x9PtWvUalKcd8iF3_WzL5vVC-90-gPeaceSaGyTNGFvrLAROj9Wbk_HyZRfoB9qsSk0ERIawXOhmdUv7q_0X2iJLgHtxHcvyjA08NNbsIlCywTb95GdZ-60noiSX7_dxBaySIe4W8sHK1NNVJE_asIqbJ86Rbjfu2u6cqur6R7KoqPTQIpCgupoTAL0mRLmjuq6IsIJRhDdw1w1qt5hdtMobkhae87o-4EJijZ8IuYbXFQ-uyoyFBeRT0TaFxQ; xq_r_token=3394242e1ca07cfa16d6f830745e7ade6b1822ac; Hm_lvt_1db88642e346389874251b5a1eded6e3=1626014204,1626050123,1626097100,1628388326; device_id=53ad799100a88e34ed535c9069a5a388; snbim_minify=true; __utma=1.623803639.1625906667.1626136094.1628388672.9; __utmc=1; Hm_lpvt_1db88642e346389874251b5a1eded6e3=1628389111; acw_tc=276082a616283938324795448e6bd032b74fd48c1908077574a5412b841337'
            }
nums=5
if __name__=="__main__":
    wb=Workbook()
    wb.active
    xueqiu=spider()
    for i in range(1,nums+1):
        #ws1=wb.create_sheet(str(i)+'_page',0)
        url="https://xueqiu.com/cubes/discover/rank/cube/list.json?category=14&page="+str(i)+"&count=20"
        json_data=xueqiu.request_data(url,header)
        time.sleep(1)
        xueqiu.json_parse(json_data)
    ws2=wb.create_sheet("end_data",0)
    for k in histroy_data_set:
        ws2.append(k)
    xueqiu.data_analysis(histroy_data_set)
    #print(data)
    ws3=wb.create_sheet("analysis_data",0)
    dict_data=[]
    for i in data:
        dict_data.append([i,data[i]])
    for j in dict_data:
        ws3.append(j)
    wb.save('data.xlsx')
